from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS
from functools import wraps
import jwt
import json
import os
import uuid
from pathlib import Path
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

app = Flask(__name__)
CORS(app)

# Configuration
app.config['SECRET_KEY'] = 'travel-app-secret-key-2024'
BASE_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = BASE_DIR.parent.parent.parent
app.config['UPLOAD_FOLDER'] = str(PROJECT_ROOT / 'uploads')

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Admin credentials (hashed)
ADMIN_CREDENTIALS = {
    'admin': generate_password_hash('admin123')
}

# Data storage paths
DATA_FILE = BASE_DIR / 'data.json'
ABOUT_FILE = BASE_DIR / 'about.json'
USERS_FILE = BASE_DIR / 'users.json'
ENQUIRIES_XLSX_FILE = BASE_DIR / 'enquiries.xlsx'
DEFAULT_DATA_SCHEMA = {
    'high_selling_packages': [],
    'all_packages': [],
    'home_images': [],
    'enquiries': []
}

def ensure_data_schema(data):
    normalized = dict(data) if isinstance(data, dict) else {}
    for key, default_value in DEFAULT_DATA_SCHEMA.items():
        if key not in normalized or not isinstance(normalized[key], list):
            normalized[key] = list(default_value)
    return normalized

# ===== UTILITY FUNCTIONS =====
def load_data():
    """Load data from JSON file with default structure"""
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            normalized = ensure_data_schema(data)
            if normalized != data:
                save_data(normalized)
            return normalized
        except Exception:
            pass
    return ensure_data_schema({})

def save_data(data):
    """Save data to JSON file"""
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(ensure_data_schema(data), f, indent=2)

def load_about():
    """Load about content"""
    if os.path.exists(ABOUT_FILE):
        try:
            with open(ABOUT_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {'content': '', 'video': ''}

def save_about(data):
    """Save about content"""
    with open(ABOUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2)

def load_users():
    if os.path.exists(USERS_FILE):
        try:
            with open(USERS_FILE, 'r', encoding='utf-8') as f:
                users = json.load(f)
            return users if isinstance(users, list) else []
        except Exception:
            return []
    return []

def save_users(users):
    with open(USERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(users, f, indent=2)

def append_enquiry_to_xlsx(enquiry):
    """Append enquiry to enquiries.xlsx in backend folder."""
    if Workbook is None or load_workbook is None:
        return 'openpyxl is not installed; enquiry saved to JSON only.'

    headers = ['ID', 'Name', 'Email', 'Contact', 'Package', 'Message', 'Timestamp']
    if ENQUIRIES_XLSX_FILE.exists():
        wb = load_workbook(ENQUIRIES_XLSX_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Enquiries'
        ws.append(headers)

    ws.append([
        enquiry.get('id', ''),
        enquiry.get('name', ''),
        enquiry.get('email', ''),
        enquiry.get('contact', ''),
        enquiry.get('package', ''),
        enquiry.get('message', ''),
        enquiry.get('timestamp', '')
    ])
    wb.save(ENQUIRIES_XLSX_FILE)
    return None

def rewrite_enquiries_xlsx(enquiries):
    """Rewrite enquiries.xlsx from the current enquiries list."""
    if Workbook is None or load_workbook is None:
        return 'openpyxl is not installed; enquiry updated in JSON only.'

    headers = ['ID', 'Name', 'Email', 'Contact', 'Package', 'Message', 'Timestamp']
    wb = Workbook()
    ws = wb.active
    ws.title = 'Enquiries'
    ws.append(headers)

    for enquiry in enquiries:
        ws.append([
            enquiry.get('id', ''),
            enquiry.get('name', ''),
            enquiry.get('email', ''),
            enquiry.get('contact', ''),
            enquiry.get('package', ''),
            enquiry.get('message', ''),
            enquiry.get('timestamp', '')
        ])

    wb.save(ENQUIRIES_XLSX_FILE)
    return None

def normalize_includes(value):
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    if isinstance(value, str):
        return [item.strip() for item in value.split(',') if item.strip()]
    return []

def token_required(f):
    """Decorator to require JWT token"""
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        
        # Check for token in Authorization header
        if 'Authorization' in request.headers:
            auth_header = request.headers['Authorization']
            try:
                token = auth_header.split(" ")[1]
            except IndexError:
                return jsonify({'message': 'Invalid token format'}), 401
        
        if not token:
            return jsonify({'message': 'Token is missing!'}), 401
        
        try:
            data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
            request.user = data
        except jwt.ExpiredSignatureError:
            return jsonify({'message': 'Token has expired'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'message': 'Invalid token'}), 401
        
        return f(*args, **kwargs)
    return decorated

# ===== AUTHENTICATION ENDPOINTS =====
@app.route('/api/auth/login', methods=['POST'])
def login():
    """Authenticate user and return JWT token"""
    data = request.json
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({'message': 'Missing username or password'}), 400
    # First try to authenticate against users.json
    users = load_users()
    matched = next((u for u in users if u.get('username') == username), None)

    if matched:
        # User exists in users.json
        if not matched.get('password') or not check_password_hash(matched['password'], password):
            return jsonify({'message': 'Invalid credentials'}), 401
        role = matched.get('role', 'user')
        user_id = matched.get('id')
    else:
        # Fallback to ADMIN_CREDENTIALS
        if username not in ADMIN_CREDENTIALS or not check_password_hash(ADMIN_CREDENTIALS[username], password):
            return jsonify({'message': 'Invalid credentials'}), 401
        role = 'admin'
        user_id = None

    # Generate JWT token (expires in 24 hours)
    token = jwt.encode({
        'username': username,
        'role': role,
        'user_id': user_id,
        'exp': datetime.utcnow() + timedelta(hours=24)
    }, app.config['SECRET_KEY'], algorithm='HS256')

    return jsonify({
        'message': 'Login successful',
        'token': token,
        'role': role
    }), 200

# ===== ENQUIRIES ENDPOINTS =====
@app.route('/api/enquiries', methods=['GET'])
@token_required
def get_enquiries():
    """Get all customer enquiries"""
    data = load_data()
    return jsonify({
        'success': True,
        'data': data.get('enquiries', [])
    }), 200

@app.route('/api/enquiries', methods=['POST'])
def create_enquiry():
    """Create a new customer enquiry (no auth required for frontend form)"""
    form_data = request.get_json(silent=True) or {}
    name = form_data.get('name', '').strip()
    email = form_data.get('email', '').strip()
    contact = form_data.get('contact', '').strip()
    package_name = form_data.get('package', '').strip()
    message = form_data.get('message', '').strip()

    if not name:
        return jsonify({'success': False, 'message': 'Name is required'}), 400
    if not email and not contact:
        return jsonify({'success': False, 'message': 'Email or contact is required'}), 400
    
    enquiry = {
        'id': str(uuid.uuid4()),
        'name': name,
        'email': email,
        'contact': contact,
        'package': package_name,
        'message': message or (f"Package enquiry for {package_name}" if package_name else ''),
        'timestamp': datetime.now().isoformat()
    }
    
    data = load_data()
    data.setdefault('enquiries', []).append(enquiry)
    save_data(data)
    xlsx_error = append_enquiry_to_xlsx(enquiry)
    
    response = {
        'success': True,
        'message': 'Enquiry created successfully',
        'enquiry': enquiry
    }
    if xlsx_error:
        response['warning'] = xlsx_error

    return jsonify(response), 201

@app.route('/api/enquiries/<enquiry_id>', methods=['DELETE'])
@token_required
def delete_enquiry(enquiry_id):
    """Delete a customer enquiry and sync enquiries.xlsx."""
    data = load_data()
    enquiries = data.get('enquiries', [])
    filtered_enquiries = [e for e in enquiries if e.get('id') != enquiry_id]

    if len(filtered_enquiries) == len(enquiries):
        return jsonify({'success': False, 'message': 'Enquiry not found'}), 404

    data['enquiries'] = filtered_enquiries
    save_data(data)
    xlsx_error = rewrite_enquiries_xlsx(filtered_enquiries)

    response = {
        'success': True,
        'message': 'Enquiry deleted successfully'
    }
    if xlsx_error:
        response['warning'] = xlsx_error

    return jsonify(response), 200

@app.route('/api/enquiries/export', methods=['GET'])
@token_required
def export_enquiries_xlsx():
    """Download enquiries.xlsx from backend storage."""
    if not ENQUIRIES_XLSX_FILE.exists():
        # Try rebuilding XLSX from current JSON enquiries when file is missing.
        data = load_data()
        xlsx_error = rewrite_enquiries_xlsx(data.get('enquiries', []))
        if xlsx_error or not ENQUIRIES_XLSX_FILE.exists():
            return jsonify({
                'success': False,
                'message': xlsx_error or 'enquiries.xlsx is not available'
            }), 404

    return send_from_directory(
        str(BASE_DIR),
        ENQUIRIES_XLSX_FILE.name,
        as_attachment=True
    )

# ===== HIGH-SELLING PACKAGES ENDPOINTS =====
@app.route('/api/high-selling-packages', methods=['GET'])
def get_high_selling_packages():
    """Get all high-selling packages"""
    data = load_data()
    return jsonify({
        'success': True,
        'data': data.get('high_selling_packages', [])
    }), 200

@app.route('/api/high-selling-packages', methods=['POST'])
@token_required
def create_high_selling_package():
    """Create a new high-selling package"""
    form_data = request.json
    
    package = {
        'id': str(uuid.uuid4()),
        'name': form_data.get('name', ''),
        'price': form_data.get('price', ''),
        'description': form_data.get('description', ''),
        'image': None,
        'created_at': datetime.now().isoformat()
    }
    
    data = load_data()
    data['high_selling_packages'].append(package)
    save_data(data)
    
    return jsonify({
        'success': True,
        'message': 'Package created successfully',
        'package': package
    }), 201

@app.route('/api/high-selling-packages/<package_id>', methods=['DELETE'])
@token_required
def delete_high_selling_package(package_id):
    """Delete a high-selling package"""
    data = load_data()
    data['high_selling_packages'] = [
        p for p in data.get('high_selling_packages', [])
        if p['id'] != package_id
    ]
    save_data(data)
    
    return jsonify({
        'success': True,
        'message': 'Package deleted successfully'
    }), 200

@app.route('/api/high-selling-packages/<package_id>', methods=['PUT'])
@token_required
def update_high_selling_package(package_id):
    """Update a high-selling package"""
    form_data = request.json
    data = load_data()
    
    for package in data.get('high_selling_packages', []):
        if package['id'] == package_id:
            package.update({
                'name': form_data.get('name', package['name']),
                'price': form_data.get('price', package['price']),
                'description': form_data.get('description', package['description'])
            })
            break
    
    save_data(data)
    return jsonify({
        'success': True,
        'message': 'Package updated successfully'
    }), 200

# ===== ALL PACKAGES ENDPOINTS =====
@app.route('/api/packages', methods=['GET'])
def get_all_packages():
    """Get all tour packages"""
    data = load_data()
    return jsonify({
        'success': True,
        'data': data.get('all_packages', [])
    }), 200

@app.route('/api/packages', methods=['POST'])
@token_required
def create_package():
    """Create a new tour package"""
    form_data = request.json
    
    package = {
        'id': str(uuid.uuid4()),
        'name': form_data.get('name', ''),
        'price': form_data.get('price', ''),
        'description': form_data.get('description', ''),
        'duration': form_data.get('duration', ''),
        'includes': normalize_includes(form_data.get('includes', [])),
        'image': None,
        'created_at': datetime.now().isoformat()
    }
    
    data = load_data()
    data['all_packages'].append(package)
    save_data(data)
    
    return jsonify({
        'success': True,
        'message': 'Package created successfully',
        'package': package
    }), 201

@app.route('/api/packages/<package_id>', methods=['DELETE'])
@token_required
def delete_package(package_id):
    """Delete a tour package"""
    data = load_data()
    data['all_packages'] = [
        p for p in data.get('all_packages', [])
        if p['id'] != package_id
    ]
    save_data(data)
    
    return jsonify({
        'success': True,
        'message': 'Package deleted successfully'
    }), 200

@app.route('/api/packages/<package_id>', methods=['PUT'])
@token_required
def update_package(package_id):
    """Update a tour package"""
    form_data = request.json
    data = load_data()
    
    for package in data.get('all_packages', []):
        if package['id'] == package_id:
            package.update({
                'name': form_data.get('name', package['name']),
                'price': form_data.get('price', package['price']),
                'description': form_data.get('description', package['description']),
                'duration': form_data.get('duration', package.get('duration', '')),
                'includes': normalize_includes(form_data.get('includes', package.get('includes', [])))
            })
            break
    
    save_data(data)
    return jsonify({
        'success': True,
        'message': 'Package updated successfully'
    }), 200

# ===== HOME IMAGES ENDPOINTS =====
@app.route('/api/home-images', methods=['GET'])
def get_home_images():
    """Get all home page images"""
    data = load_data()
    return jsonify({
        'success': True,
        'data': data.get('home_images', [])
    }), 200

@app.route('/api/home-images', methods=['POST'])
@token_required
def upload_home_image():
    """Upload a home page image"""
    if 'file' not in request.files:
        return jsonify({'message': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'message': 'No file selected'}), 400
    
    try:
        filename = f"{uuid.uuid4()}_{file.filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        image = {
            'id': str(uuid.uuid4()),
            'url': f'/uploads/{filename}',
            'filename': filename,
            'uploaded_at': datetime.now().isoformat()
        }
        
        data = load_data()
        data['home_images'].append(image)
        save_data(data)
        
        return jsonify({
            'success': True,
            'message': 'Image uploaded successfully',
            'image': image
        }), 201
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@app.route('/api/home-images/<image_id>', methods=['DELETE'])
@token_required
def delete_home_image(image_id):
    """Delete a home page image"""
    data = load_data()
    
    # Find and delete the image
    image_to_delete = None
    for img in data.get('home_images', []):
        if img['id'] == image_id:
            image_to_delete = img
            break
    
    if image_to_delete:
        try:
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], image_to_delete.get('filename', ''))
            if os.path.exists(filepath):
                os.remove(filepath)
        except:
            pass
        
        data['home_images'] = [
            img for img in data.get('home_images', [])
            if img['id'] != image_id
        ]
        save_data(data)
    
    return jsonify({
        'success': True,
        'message': 'Image deleted successfully'
    }), 200

@app.route('/api/packages/<package_id>/image', methods=['POST'])
@token_required
def upload_package_image(package_id):
    """Upload image for a package"""
    if 'file' not in request.files:
        return jsonify({'message': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'message': 'No file selected'}), 400
    
    try:
        filename = f"{uuid.uuid4()}_{file.filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        image_url = f'/uploads/{filename}'
        
        # Update package with image
        data = load_data()
        for pkg in data.get('all_packages', []):
            if pkg['id'] == package_id:
                pkg['image'] = image_url
                break
        
        save_data(data)
        
        return jsonify({
            'success': True,
            'message': 'Image uploaded successfully',
            'image': image_url
        }), 200
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@app.route('/api/high-selling-packages/<package_id>/image', methods=['POST'])
@token_required
def upload_high_selling_package_image(package_id):
    """Upload image for a high-selling package"""
    if 'file' not in request.files:
        return jsonify({'message': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'message': 'No file selected'}), 400
    
    try:
        filename = f"{uuid.uuid4()}_{file.filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        image_url = f'/uploads/{filename}'
        
        # Update package with image
        data = load_data()
        for pkg in data.get('high_selling_packages', []):
            if pkg['id'] == package_id:
                pkg['image'] = image_url
                break
        
        save_data(data)
        
        return jsonify({
            'success': True,
            'message': 'Image uploaded successfully',
            'image': image_url
        }), 200
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

# ===== ABOUT/HISTORY ENDPOINTS =====
@app.route('/api/about', methods=['GET'])
def get_about():
    """Get about content"""
    about_data = load_about()
    return jsonify({
        'success': True,
        'data': about_data
    }), 200

@app.route('/api/about', methods=['PUT'])
@token_required
def update_about():
    """Update about content and video"""
    form_data = request.json
    
    about_data = {
        'content': form_data.get('content', ''),
        'video': form_data.get('video', ''),
        'updated_at': datetime.now().isoformat()
    }
    
    save_about(about_data)
    
    return jsonify({
        'success': True,
        'message': 'About content updated successfully'
    }), 200

@app.route('/api/about/video', methods=['POST'])
@token_required
def upload_about_video():
    """Upload video for about page"""
    if 'file' not in request.files:
        return jsonify({'message': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'message': 'No file selected'}), 400
    
    try:
        filename = f"{uuid.uuid4()}_{file.filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        video_url = f'/uploads/{filename}'
        
        # Update about with video URL
        about_data = load_about()
        about_data['video'] = video_url
        about_data['updated_at'] = datetime.now().isoformat()
        save_about(about_data)
        
        return jsonify({
            'success': True,
            'message': 'Video uploaded successfully',
            'data': {
                'video': video_url,
                'url': video_url
            }
        }), 200
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

# ===== USER MANAGEMENT ENDPOINTS =====

@app.route('/api/users', methods=['GET'])
@token_required
def get_users():
    """Get all users"""
    users = load_users()
    response_users = [
        {
            'id': u.get('id'),
            'username': u.get('username', ''),
            'email': u.get('email', ''),
            'role': u.get('role', 'user')
        }
        for u in users
    ]
    return jsonify({'success': True, 'data': response_users}), 200

@app.route('/api/users', methods=['POST'])
@token_required
def create_user():
    """Create a new user"""
    data = request.get_json()
    username = data.get('username')
    email = data.get('email')
    password = data.get('password')
    role = data.get('role', 'user')
    
    if not username or not email or not password:
        return jsonify({'success': False, 'message': 'Missing required fields'}), 400
    
    users = load_users()
    
    # Check if user already exists
    if any(u['username'] == username for u in users):
        return jsonify({'success': False, 'message': 'Username already exists'}), 400
    
    if any(u['email'] == email for u in users):
        return jsonify({'success': False, 'message': 'Email already exists'}), 400
    
    new_user = {
        'id': str(uuid.uuid4()),
        'username': username,
        'email': email,
        'password': generate_password_hash(password),
        'role': role,
        'created_at': datetime.now().isoformat()
    }
    
    users.append(new_user)
    save_users(users)
    
    # Return user without password
    response_user = {'id': new_user['id'], 'username': new_user['username'], 'email': new_user['email'], 'role': new_user['role']}
    return jsonify({'success': True, 'data': response_user}), 201

@app.route('/api/users/<user_id>', methods=['PUT'])
@token_required
def update_user(user_id):
    """Update an existing user"""
    data = request.get_json()
    users = load_users()
    
    user_index = next((i for i, u in enumerate(users) if u['id'] == user_id), None)
    if user_index is None:
        return jsonify({'success': False, 'message': 'User not found'}), 404
    
    user = users[user_index]
    
    # Update username if provided and unique
    if 'username' in data and data['username'] != user['username']:
        if any(u['username'] == data['username'] for u in users):
            return jsonify({'success': False, 'message': 'Username already exists'}), 400
        user['username'] = data['username']
    
    # Update email if provided and unique
    if 'email' in data and data['email'] != user['email']:
        if any(u['email'] == data['email'] for u in users):
            return jsonify({'success': False, 'message': 'Email already exists'}), 400
        user['email'] = data['email']
    
    # Update password if provided
    if 'password' in data and data['password']:
        user['password'] = generate_password_hash(data['password'])
    
    # Update role if provided
    if 'role' in data:
        user['role'] = data['role']
    
    user['updated_at'] = datetime.now().isoformat()
    users[user_index] = user
    save_users(users)
    
    # Return user without password
    response_user = {'id': user['id'], 'username': user['username'], 'email': user['email'], 'role': user['role']}
    return jsonify({'success': True, 'data': response_user})

@app.route('/api/users/<user_id>', methods=['DELETE'])
@token_required
def delete_user(user_id):
    """Delete a user"""
    users = load_users()
    
    user_index = next((i for i, u in enumerate(users) if u['id'] == user_id), None)
    if user_index is None:
        return jsonify({'success': False, 'message': 'User not found'}), 404
    
    users.pop(user_index)
    save_users(users)
    
    return jsonify({'success': True, 'message': 'User deleted successfully'})

# ===== STATIC FILE SERVING =====
@app.route('/uploads/<filename>')
def serve_upload(filename):
    """Serve uploaded files"""
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# ===== ERROR HANDLERS =====
@app.errorhandler(404)
def not_found(error):
    return jsonify({'message': 'Resource not found'}), 404

@app.errorhandler(500)
def server_error(error):
    return jsonify({'message': 'Internal server error'}), 500

if __name__ == '__main__':
    app.run(debug=False, port=5000, host='0.0.0.0')
